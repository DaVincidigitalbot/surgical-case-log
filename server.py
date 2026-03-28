#!/usr/bin/env python3
"""
Clinical Case Log — Backend API
PostgreSQL database for persistent storage with user accounts.
"""
from flask import Flask, request, jsonify, send_from_directory, Response
from flask_cors import CORS
import os
import hashlib
import secrets
import json
from datetime import datetime
import psycopg2
from psycopg2.extras import RealDictCursor
import urllib.parse
import io
import smtplib
import random
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import timedelta

app = Flask(__name__, static_folder='.', static_url_path='')
CORS(app)

# PostgreSQL connection
DB_PASSWORD = urllib.parse.quote(os.environ.get('DB_PASSWORD', 'CaseLog2026!'))
DATABASE_URL = os.environ.get('DATABASE_URL', 
    f"postgresql://postgres.qooenxfumvitimfbuocp:{DB_PASSWORD}@aws-1-us-east-1.pooler.supabase.com:5432/postgres")

def get_db():
    conn = psycopg2.connect(DATABASE_URL)
    return conn

def init_db():
    """Initialize database tables if they don't exist."""
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                email TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                name TEXT,
                token TEXT UNIQUE,
                plan TEXT DEFAULT 'trial',
                license_key TEXT,
                plan_expires_at TEXT,
                program_type TEXT DEFAULT 'medical',
                trial_started_at TIMESTAMP,
                trial_ends_at TIMESTAMP,
                trial_case_limit INTEGER DEFAULT 25,
                trial_status TEXT DEFAULT 'active',
                trial_end_reason TEXT,
                created_at TIMESTAMP DEFAULT NOW()
            )
        ''')
        # Add trial columns to existing tables (safe to run repeatedly)
        for col, coltype, default in [
            ('trial_started_at', 'TIMESTAMP', None),
            ('trial_ends_at', 'TIMESTAMP', None),
            ('trial_case_limit', 'INTEGER', '25'),
            ('trial_status', 'TEXT', "'active'"),
            ('trial_end_reason', 'TEXT', None),
            ('trial_completed_at', 'TIMESTAMP', None),
            ('email_sequence_status', 'TEXT', "'pending'"),
            ('last_email_sent_at', 'TIMESTAMP', None),
            ('emails_sent_count', 'INTEGER', '0'),
        ]:
            try:
                default_clause = f" DEFAULT {default}" if default else ""
                cur.execute(f"ALTER TABLE users ADD COLUMN IF NOT EXISTS {col} {coltype}{default_clause}")
            except:
                conn.rollback()
        
        # Email log table
        cur.execute('''
            CREATE TABLE IF NOT EXISTS email_log (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL REFERENCES users(id),
                email_type TEXT NOT NULL,
                sequence_number INTEGER,
                subject TEXT,
                sent_at TIMESTAMP DEFAULT NOW(),
                status TEXT DEFAULT 'sent'
            )
        ''')
        try:
            cur.execute('CREATE INDEX IF NOT EXISTS idx_email_log_user ON email_log(user_id)')
        except:
            conn.rollback()
        cur.execute('''
            CREATE TABLE IF NOT EXISTS cases (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL REFERENCES users(id),
                date TEXT,
                age INTEGER,
                sex TEXT,
                rotation TEXT,
                procedure_name TEXT,
                cpt_code TEXT,
                role TEXT,
                approach TEXT,
                attending TEXT,
                complications TEXT DEFAULT 'None',
                ebl INTEGER,
                or_time INTEGER,
                notes TEXT,
                diagnosis TEXT,
                created_at TIMESTAMP DEFAULT NOW(),
                updated_at TIMESTAMP DEFAULT NOW()
            )
        ''')
        cur.execute('''
            CREATE TABLE IF NOT EXISTS milestones (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL REFERENCES users(id),
                category TEXT NOT NULL,
                count INTEGER DEFAULT 0,
                UNIQUE(user_id, category)
            )
        ''')
        cur.execute('''
            CREATE TABLE IF NOT EXISTS user_attendings (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL REFERENCES users(id),
                name TEXT NOT NULL,
                specialty TEXT,
                UNIQUE(user_id, name)
            )
        ''')
        # Add program_type column to users if not exists
        cur.execute('''
            DO $$ BEGIN
                ALTER TABLE users ADD COLUMN program_type TEXT DEFAULT 'medical';
            EXCEPTION WHEN duplicate_column THEN NULL;
            END $$
        ''')
        # Add case_id column (ACGME-required unique case identifier)
        cur.execute('''
            DO $$ BEGIN
                ALTER TABLE cases ADD COLUMN case_id TEXT;
            EXCEPTION WHEN duplicate_column THEN NULL;
            END $$
        ''')
        cur.execute('CREATE INDEX IF NOT EXISTS idx_cases_case_id ON cases(user_id, case_id)')
        # RNFA cases table
        cur.execute('''
            CREATE TABLE IF NOT EXISTS rnfa_cases (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL REFERENCES users(id),
                case_number INTEGER,
                procedure_name TEXT,
                surgeon_name TEXT,
                case_date TEXT,
                case_time TEXT,
                preop_time INTEGER,
                intraop_time INTEGER,
                postop_time INTEGER,
                preop_behaviors TEXT,
                draping_prepping TEXT,
                special_instruments TEXT,
                medical_devices TEXT,
                exposure_hemostasis TEXT,
                layer_closure TEXT,
                postop_behaviors TEXT,
                learning_log TEXT,
                created_at TIMESTAMP DEFAULT NOW()
            )
        ''')
        # Password reset table
        cur.execute('''
            CREATE TABLE IF NOT EXISTS password_resets (
                id SERIAL PRIMARY KEY,
                email TEXT NOT NULL,
                code TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                used BOOLEAN DEFAULT FALSE
            )
        ''')
        cur.execute('CREATE INDEX IF NOT EXISTS idx_cases_user ON cases(user_id)')
        cur.execute('CREATE INDEX IF NOT EXISTS idx_cases_date ON cases(user_id, date)')
        cur.execute('CREATE INDEX IF NOT EXISTS idx_user_attendings_user ON user_attendings(user_id)')
        cur.execute('CREATE INDEX IF NOT EXISTS idx_rnfa_cases_user ON rnfa_cases(user_id)')
        cur.execute('CREATE INDEX IF NOT EXISTS idx_password_resets_email ON password_resets(email)')
        conn.commit()
        conn.close()
        print("[INIT] PostgreSQL database initialized successfully", flush=True)
    except Exception as e:
        print(f"[INIT] DB init error: {e}", flush=True)

init_db()

def hash_password(password):
    salt = secrets.token_hex(16)
    h = hashlib.pbkdf2_hmac('sha256', password.encode(), salt.encode(), 100000)
    return salt + ':' + h.hex()

def verify_password(stored, provided):
    salt, stored_hash = stored.split(':')
    h = hashlib.pbkdf2_hmac('sha256', provided.encode(), salt.encode(), 100000)
    return h.hex() == stored_hash

def check_trial_status(user):
    """Check if a trial user can still log cases. Returns (can_log, info_dict).
    Paid users always return True. Trial users checked against 25 cases / 7 days."""
    plan = user.get('plan', 'trial')
    
    # Paid users — no restrictions
    if plan in ('subscription', 'monthly', 'annual', '90day', 'pro'):
        return True, {'plan': plan, 'active': True}
    
    # Demo account — no restrictions
    if plan == 'demo':
        return True, {'plan': 'demo', 'active': True}
    
    # Expired plans
    if plan == 'expired':
        return False, {'plan': 'expired', 'active': False, 'reason': 'subscription_expired'}
    
    # Trial users — enforce limits
    trial_status = user.get('trial_status', 'active')
    
    # Already marked expired
    if trial_status in ('expired', 'converted'):
        return trial_status == 'converted', {
            'plan': 'trial',
            'trial_status': trial_status,
            'active': trial_status == 'converted',
            'reason': user.get('trial_end_reason', 'unknown')
        }
    
    # Check time limit
    trial_ends_at = user.get('trial_ends_at')
    trial_started_at = user.get('trial_started_at')
    from datetime import datetime, timezone, timedelta
    now = datetime.now(timezone.utc)
    
    # Backfill: if trial user has no trial_ends_at, set from created_at
    if not trial_ends_at and plan == 'trial':
        created = user.get('created_at')
        if created:
            if isinstance(created, str):
                created = datetime.fromisoformat(created.replace('Z', '+00:00'))
            if created.tzinfo is None:
                created = created.replace(tzinfo=timezone.utc)
            trial_ends_at = created + timedelta(days=7)
            trial_started_at = created
            # Persist the backfill
            conn2 = get_db()
            cur2 = conn2.cursor()
            cur2.execute('UPDATE users SET trial_started_at = %s, trial_ends_at = %s WHERE id = %s AND trial_ends_at IS NULL',
                        (trial_started_at, trial_ends_at, user['id']))
            conn2.commit()
            conn2.close()
    time_expired = False
    days_remaining = 7
    
    if trial_ends_at:
        if isinstance(trial_ends_at, str):
            trial_ends_at = datetime.fromisoformat(trial_ends_at.replace('Z', '+00:00'))
        if trial_ends_at.tzinfo is None:
            trial_ends_at = trial_ends_at.replace(tzinfo=timezone.utc)
        time_expired = now >= trial_ends_at
        days_remaining = max(0, (trial_ends_at - now).days)
    
    # Count cases (both medical + RNFA)
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute('SELECT COUNT(*) as cnt FROM cases WHERE user_id = %s', (user['id'],))
    medical_count = cur.fetchone()['cnt']
    cur.execute('SELECT COUNT(*) as cnt FROM rnfa_cases WHERE user_id = %s', (user['id'],))
    rnfa_count = cur.fetchone()['cnt']
    total_cases = medical_count + rnfa_count
    
    case_limit = user.get('trial_case_limit') or 25
    case_limit_hit = total_cases >= case_limit
    
    # Determine if trial should expire
    if case_limit_hit or time_expired:
        # Expire the trial
        reason = 'case_limit' if case_limit_hit else 'time_limit'
        # If both hit simultaneously, prioritize case_limit
        if case_limit_hit and time_expired:
            reason = 'case_limit'
        
        cur.execute('''UPDATE users SET trial_status = 'expired', trial_end_reason = %s,
                      trial_completed_at = NOW(), email_sequence_status = 'active'
                      WHERE id = %s AND trial_status = 'active' ''', (reason, user['id']))
        conn.commit()
        conn.close()
        
        return False, {
            'plan': 'trial',
            'trial_status': 'expired',
            'active': False,
            'reason': reason,
            'cases_used': total_cases,
            'case_limit': case_limit,
            'days_remaining': 0
        }
    
    conn.close()
    
    return True, {
        'plan': 'trial',
        'trial_status': 'active',
        'active': True,
        'cases_used': total_cases,
        'case_limit': case_limit,
        'cases_remaining': case_limit - total_cases,
        'days_remaining': days_remaining
    }


def get_user_from_token(token):
    if not token:
        return None
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute('SELECT * FROM users WHERE token = %s', (token,))
        user = cur.fetchone()
        conn.close()
        return user
    except:
        return None

def auth_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        user = get_user_from_token(token)
        if not user:
            return jsonify({'error': 'Unauthorized'}), 401
        request.user = user
        return f(*args, **kwargs)
    return decorated

# ============ AUTH ============

@app.route('/api/register', methods=['POST'])
def register():
    data = request.json
    email = data.get('email', '').strip().lower()
    password = data.get('password', '')
    name = data.get('name', '')
    
    if not email or not password:
        return jsonify({'error': 'Email and password required'}), 400
    if len(password) < 6:
        return jsonify({'error': 'Password must be 6+ characters'}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute('SELECT id FROM users WHERE email = %s', (email,))
    existing = cur.fetchone()
    if existing:
        conn.close()
        return jsonify({'error': 'Email already registered'}), 409
    
    token = secrets.token_urlsafe(32)
    pw_hash = hash_password(password)
    
    cur.execute('''INSERT INTO users (email, password_hash, name, token, plan, trial_started_at, trial_ends_at, trial_case_limit, trial_status) 
                 VALUES (%s, %s, %s, %s, 'trial', NOW(), NOW() + INTERVAL '7 days', 25, 'active')''',
                 (email, pw_hash, name, token))
    conn.commit()
    conn.close()
    
    return jsonify({'token': token, 'name': name, 'email': email, 'plan': 'trial', 'plan_expires_at': None, 'program_type': 'medical',
                    'trial_status': 'active', 'trial_case_limit': 25, 'trial_cases_used': 0, 'trial_days_remaining': 7})

@app.route('/api/login', methods=['POST'])
def login():
    data = request.json
    email = data.get('email', '').strip().lower()
    password = data.get('password', '')
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute('SELECT * FROM users WHERE email = %s', (email,))
    user = cur.fetchone()
    
    if not user or not verify_password(user['password_hash'], password):
        conn.close()
        return jsonify({'error': 'Invalid email or password'}), 401
    
    # Reuse existing token if present, otherwise generate new one
    token = user['token']
    if not token:
        token = secrets.token_urlsafe(32)
        cur.execute('UPDATE users SET token = %s WHERE id = %s', (token, user['id']))
        conn.commit()
    conn.close()
    
    return jsonify({'token': token, 'name': user['name'], 'email': user['email'], 'plan': user['plan'] or 'trial', 'plan_expires_at': user['plan_expires_at'], 'program_type': user.get('program_type', 'medical') or 'medical'})

# ============ PASSWORD RESET ============

def send_reset_email(email, code):
    """Send password reset email via SMTP"""
    try:
        smtp_email = "Graydon.F.Stallard@gmail.com"
        smtp_password = "dlrypwmlbmualsxv"
        
        msg = MIMEMultipart()
        msg['From'] = "Clinical Case Log <Graydon.F.Stallard@gmail.com>"
        msg['To'] = email
        msg['Subject'] = "Clinical Case Log — Password Reset Code"
        
        body = f"""
Hi there,

You requested a password reset for your Clinical Case Log account.

Your 6-digit reset code is: {code}

This code expires in 15 minutes.

If you didn't request this reset, you can safely ignore this email.

Best regards,
Clinical Case Log Support Team
support@clinicalcaselog.com
        """
        
        msg.attach(MIMEText(body, 'plain'))
        
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(smtp_email, smtp_password)
        server.send_message(msg)
        server.quit()
        
        return True
    except Exception as e:
        print(f"Email send error: {e}", flush=True)
        return False


# ============ TRIAL CONVERSION EMAIL SYSTEM ============

TRIAL_EMAIL_SEQUENCE = [
    {
        'number': 1,
        'delay_hours': 0,
        'subject': 'Your cases are saved — continue where you left off',
        'body': """Hi {name},

You've logged 25 cases in Clinical Case Log Pro — that's real progress toward organized surgical training documentation.

Your trial has reached its limit, but every case you logged is saved and waiting for you.

Upgrade now to continue logging cases, track ACGME milestones, and export your complete case log to Excel or PDF anytime.

Upgrade for $49.99/year: {upgrade_url}

Your data isn't going anywhere. Pick up right where you left off.

--
Clinical Case Log Pro
Built by a practicing surgeon
clinicalcaselog.com"""
    },
    {
        'number': 2,
        'delay_hours': 24,
        'subject': "Don't lose momentum on your case log",
        'body': """Hi {name},

You were building something valuable — 25 documented surgical cases with CPT codes, attendings, and procedure details all in one place.

Most residents and students spend hours reconstructing their case logs before evaluations. You were doing it in 30 seconds per case.

That efficiency is still here. Upgrade and keep logging.

Continue logging cases: {upgrade_url}

$49.99/year. Unlimited cases. Full ACGME tracking. Excel + PDF export.

--
Clinical Case Log Pro
clinicalcaselog.com"""
    },
    {
        'number': 3,
        'delay_hours': 72,
        'subject': 'Quick note about your case log',
        'body': """Hi {name},

Just checking in. Your 25 logged cases are still saved in Clinical Case Log Pro.

The students and residents who stay on top of their case documentation consistently have smoother evaluations and stronger applications. The ones who fall behind spend entire weekends trying to reconstruct months of cases from memory.

You've already built the habit. Don't let it lapse.

Upgrade now: {upgrade_url}

--
Clinical Case Log Pro
clinicalcaselog.com"""
    },
    {
        'number': 4,
        'delay_hours': 168,
        'subject': 'Your case log is still here',
        'body': """Hi {name},

This is the last note from us — your 25 cases in Clinical Case Log Pro are still saved and accessible.

If you're ready to continue logging, your account is waiting:

Upgrade for $49.99/year: {upgrade_url}

If the timing isn't right, no worries. Your data will be here whenever you're ready.

--
Clinical Case Log Pro
clinicalcaselog.com"""
    }
]


def send_conversion_email(user_email, user_name, sequence_number, user_id):
    """Send a specific email from the trial conversion sequence."""
    if sequence_number < 1 or sequence_number > len(TRIAL_EMAIL_SEQUENCE):
        return False
    
    template = TRIAL_EMAIL_SEQUENCE[sequence_number - 1]
    upgrade_url = 'https://clinicalcaselog.com/login?tab=register&upgrade=true'
    
    name = user_name or 'there'
    subject = template['subject']
    body = template['body'].format(name=name, upgrade_url=upgrade_url)
    
    try:
        smtp_email = "Graydon.F.Stallard@gmail.com"
        smtp_password = "dlrypwmlbmualsxv"
        
        msg = MIMEMultipart()
        msg['From'] = "Clinical Case Log <Graydon.F.Stallard@gmail.com>"
        msg['To'] = user_email
        msg['Subject'] = subject
        msg['Reply-To'] = "support@clinicalcaselog.com"
        msg['List-Unsubscribe'] = '<mailto:support@clinicalcaselog.com?subject=unsubscribe>'
        
        msg.attach(MIMEText(body, 'plain'))
        
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(smtp_email, smtp_password)
        server.send_message(msg)
        server.quit()
        
        # Log the email
        try:
            conn = get_db()
            cur = conn.cursor()
            cur.execute('''INSERT INTO email_log (user_id, email_type, sequence_number, subject)
                          VALUES (%s, %s, %s, %s)''',
                       (user_id, 'trial_conversion', sequence_number, subject))
            cur.execute('''UPDATE users SET last_email_sent_at = NOW(), emails_sent_count = COALESCE(emails_sent_count, 0) + 1
                          WHERE id = %s''', (user_id,))
            conn.commit()
            conn.close()
        except Exception as e:
            print(f"[EMAIL] Log error: {e}", flush=True)
        
        print(f"[EMAIL] Sent conversion email #{sequence_number} to {user_email}", flush=True)
        return True
    except Exception as e:
        print(f"[EMAIL] Send error for {user_email}: {e}", flush=True)
        return False


@app.route('/api/internal/process-trial-emails', methods=['POST'])
def process_trial_emails():
    """Process trial conversion email queue. Called by external cron hourly."""
    auth_key = request.headers.get('X-Internal-Key', '')
    if auth_key != 'stallard2026internal':
        return jsonify({'error': 'Unauthorized'}), 401
    
    from datetime import timezone
    now = datetime.now(timezone.utc)
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    # Find users with expired trials who need emails
    cur.execute('''
        SELECT id, email, name, trial_completed_at, emails_sent_count, last_email_sent_at,
               email_sequence_status, trial_end_reason
        FROM users 
        WHERE trial_status = 'expired' 
          AND email_sequence_status = 'active'
          AND COALESCE(emails_sent_count, 0) < 4
    ''')
    users = cur.fetchall()
    
    sent_count = 0
    skipped_count = 0
    
    for user in users:
        user_id = user['id']
        emails_sent = user.get('emails_sent_count') or 0
        trial_completed = user.get('trial_completed_at')
        
        if not trial_completed:
            continue
        
        if isinstance(trial_completed, str):
            trial_completed = datetime.fromisoformat(trial_completed.replace('Z', '+00:00'))
        if trial_completed.tzinfo is None:
            trial_completed = trial_completed.replace(tzinfo=timezone.utc)
        
        next_email_num = emails_sent + 1
        if next_email_num > 4:
            cur.execute("UPDATE users SET email_sequence_status = 'completed' WHERE id = %s", (user_id,))
            continue
        
        # Check timing
        template = TRIAL_EMAIL_SEQUENCE[next_email_num - 1]
        send_after = trial_completed + timedelta(hours=template['delay_hours'])
        
        if now < send_after:
            skipped_count += 1
            continue
        
        # Prevent duplicates
        cur.execute('''SELECT id FROM email_log 
                      WHERE user_id = %s AND email_type = 'trial_conversion' AND sequence_number = %s''',
                   (user_id, next_email_num))
        if cur.fetchone():
            skipped_count += 1
            continue
        
        # Send
        success = send_conversion_email(user['email'], user['name'], next_email_num, user_id)
        if success:
            sent_count += 1
        
        if next_email_num >= 4:
            cur.execute("UPDATE users SET email_sequence_status = 'completed' WHERE id = %s", (user_id,))
    
    conn.commit()
    conn.close()
    
    return jsonify({
        'processed': len(users),
        'sent': sent_count,
        'skipped': skipped_count,
        'timestamp': now.isoformat()
    })


@app.route('/api/forgot-password', methods=['POST'])
def forgot_password():
    data = request.json
    email = data.get('email', '').strip().lower()
    
    if not email:
        return jsonify({'error': 'Email is required'}), 400
    
    # Check if user exists
    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT id FROM users WHERE email = %s', (email,))
    user = cur.fetchone()
    
    if not user:
        # Still return success to avoid email enumeration
        return jsonify({'success': True, 'message': 'If an account with this email exists, a reset code has been sent.'})
    
    # Generate 6-digit code
    code = str(random.randint(100000, 999999))
    
    # Store reset code (expire old codes first)
    cur.execute('UPDATE password_resets SET used = TRUE WHERE email = %s AND used = FALSE', (email,))
    cur.execute('INSERT INTO password_resets (email, code) VALUES (%s, %s)', (email, code))
    conn.commit()
    conn.close()
    
    # Send email
    if send_reset_email(email, code):
        return jsonify({'success': True, 'message': 'Reset code sent to your email.'})
    else:
        return jsonify({'error': 'Failed to send reset email. Please try again.'}), 500

@app.route('/api/reset-password', methods=['POST'])
def reset_password():
    data = request.json
    email = data.get('email', '').strip().lower()
    code = data.get('code', '').strip()
    new_password = data.get('new_password', '')
    
    if not email or not code or not new_password:
        return jsonify({'error': 'Email, code, and new password are required'}), 400
    
    if len(new_password) < 6:
        return jsonify({'error': 'Password must be at least 6 characters'}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    # Verify reset code
    cur.execute('''
        SELECT id FROM password_resets 
        WHERE email = %s AND code = %s AND used = FALSE 
        AND created_at > NOW() - INTERVAL '15 minutes'
    ''', (email, code))
    reset_record = cur.fetchone()
    
    if not reset_record:
        conn.close()
        return jsonify({'error': 'Invalid or expired reset code'}), 400
    
    # Check if user exists
    cur.execute('SELECT id FROM users WHERE email = %s', (email,))
    user = cur.fetchone()
    
    if not user:
        conn.close()
        return jsonify({'error': 'User not found'}), 404
    
    # Update password
    pw_hash = hash_password(new_password)
    cur.execute('UPDATE users SET password_hash = %s WHERE email = %s', (pw_hash, email))
    
    # Mark reset code as used
    cur.execute('UPDATE password_resets SET used = TRUE WHERE id = %s', (reset_record['id'],))
    
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'message': 'Password updated successfully'})

# ============ TOKEN CHECK ============

@app.route('/api/me', methods=['GET'])
def check_token():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    user = get_user_from_token(token)
    if not user:
        return jsonify({'error': 'Invalid token'}), 401
    return jsonify({'name': user['name'], 'email': user['email'], 'plan': user['plan'] or 'trial', 'program_type': user.get('program_type', 'medical') or 'medical'})

# ============ LICENSE VERIFICATION ============

@app.route('/api/activate', methods=['POST'])
@auth_required
def activate_by_email():
    import requests as req
    user_email = request.user['email']

    # 1) Check Stripe first
    try:
        customers = stripe_lib.Customer.list(email=user_email, limit=5)
        for customer in customers.auto_paging_iter():
            subs = stripe_lib.Subscription.list(customer=customer.id, status='active', limit=1)
            if not subs.data:
                subs = stripe_lib.Subscription.list(customer=customer.id, status='trialing', limit=1)
            if subs.data:
                conn = get_db()
                cur = conn.cursor()
                sub = subs.data[0]
                expires = datetime.utcfromtimestamp(sub.current_period_end).strftime('%Y-%m-%d %H:%M:%S')
                cur.execute('''UPDATE users SET plan = %s, plan_expires_at = %s, trial_status = 'converted',
                           email_sequence_status = 'stopped_upgrade' WHERE id = %s''',
                           ('subscription', expires, request.user['id']))
                conn.commit()
                conn.close()
                return jsonify({'success': True, 'plan': 'subscription', 'plan_expires_at': expires})
    except Exception as e:
        print(f"Stripe activation check error: {e}")

    # 2) Fallback to Gumroad for legacy purchases
    GUMROAD_TOKEN = "X_X3q6Cw7bJuRkvmwKRGIwhN9Comr3wkjKkMXrv-ZFE"
    
    try:
        r = req.get(f'https://api.gumroad.com/v2/sales?access_token={GUMROAD_TOKEN}&email={user_email}', timeout=15)
        if r.status_code == 200:
            result = r.json()
            sales = result.get('sales', [])
            if not sales:
                return jsonify({'error': 'No active subscription found.', 'needs_purchase': True}), 404
            
            conn = get_db()
            cur = conn.cursor()
            for sale in sales:
                if sale.get('refunded') or sale.get('disputed'):
                    continue
                is_subscription = sale.get('subscription_id') is not None
                recurrence = sale.get('recurrence')
                if is_subscription or recurrence:
                    cancelled = sale.get('cancelled', False)
                    ended = sale.get('ended', False)
                    if not cancelled and not ended:
                        plan = 'subscription'
                        expires = None
                    else:
                        continue
                else:
                    plan = 'annual'
                    from datetime import timedelta
                    purchase_date = sale.get('created_at', '')
                    try:
                        pd = datetime.fromisoformat(purchase_date.replace('Z', '+00:00'))
                        expires = (pd + timedelta(days=365)).strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        expires = (datetime.utcnow() + timedelta(days=365)).strftime('%Y-%m-%d %H:%M:%S')
                
                cur.execute('''UPDATE users SET plan = %s, plan_expires_at = %s, trial_status = 'converted',
                           email_sequence_status = 'stopped_upgrade' WHERE id = %s''',
                           (plan, expires, request.user['id']))
                conn.commit()
                conn.close()
                return jsonify({'success': True, 'plan': plan, 'plan_expires_at': expires})
            
            conn.close()
            return jsonify({'error': 'No active purchase found.', 'needs_purchase': True}), 404
        else:
            return jsonify({'error': 'Could not verify purchase'}), 400
    except Exception as e:
        return jsonify({'error': f'Verification failed: {str(e)}'}), 500

@app.route('/api/plan', methods=['GET'])
@auth_required
def get_plan():
    plan = request.user['plan'] or 'trial'
    expires = request.user['plan_expires_at']
    if plan in ('annual', '90day') and expires:
        if datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S') > expires:
            plan = 'expired'
    
    # Include trial info for trial users
    if plan == 'trial':
        can_log, trial_info = check_trial_status(request.user)
        return jsonify({
            'plan': 'trial',
            'plan_expires_at': expires,
            'active': can_log,
            'trial_status': trial_info.get('trial_status', 'active'),
            'cases_used': trial_info.get('cases_used', 0),
            'case_limit': trial_info.get('case_limit', 25),
            'cases_remaining': trial_info.get('cases_remaining', 25),
            'days_remaining': trial_info.get('days_remaining', 7),
            'reason': trial_info.get('reason')
        })
    
    return jsonify({'plan': plan, 'plan_expires_at': expires, 'active': plan in ('monthly', 'annual', '90day', 'subscription', 'pro')})

@app.route('/api/trial-status', methods=['GET'])
@auth_required
def trial_status():
    """Return detailed trial status for frontend display."""
    can_log, info = check_trial_status(request.user)
    return jsonify(info)

# ============ CASES ============

@app.route('/api/cases', methods=['GET'])
@auth_required
def get_cases():
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute('SELECT * FROM cases WHERE user_id = %s ORDER BY date DESC, id DESC', (request.user['id'],))
    cases = cur.fetchall()
    conn.close()
    # Convert datetime objects to strings
    for c in cases:
        if c.get('created_at') and hasattr(c['created_at'], 'isoformat'):
            c['created_at'] = c['created_at'].isoformat()
        if c.get('updated_at') and hasattr(c['updated_at'], 'isoformat'):
            c['updated_at'] = c['updated_at'].isoformat()
    return jsonify(cases)

@app.route('/api/cases/mark-exported', methods=['POST'])
@auth_required
def mark_cases_exported():
    data = request.get_json() or {}
    case_ids = data.get('caseIds', [])
    if not case_ids:
        return jsonify({'error': 'No case IDs provided'}), 400
    try:
        conn = get_db()
        cur = conn.cursor()
        # Add acgme_exported column if it doesn't exist
        try:
            cur.execute("ALTER TABLE cases ADD COLUMN IF NOT EXISTS acgme_exported BOOLEAN DEFAULT FALSE")
            cur.execute("ALTER TABLE cases ADD COLUMN IF NOT EXISTS acgme_exported_at TIMESTAMP")
            conn.commit()
        except:
            conn.rollback()
        
        placeholders = ','.join(['%s'] * len(case_ids))
        cur.execute(
            f"UPDATE cases SET acgme_exported = TRUE, acgme_exported_at = NOW() WHERE id IN ({placeholders}) AND user_id = %s",
            (*case_ids, request.user['id'])
        )
        updated = cur.rowcount
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'updated': updated})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/cases', methods=['POST'])
@auth_required
def add_case():
    # TRIAL ENFORCEMENT — check before allowing case creation
    can_log, trial_info = check_trial_status(request.user)
    if not can_log:
        return jsonify({
            'error': 'Trial limit reached. Upgrade to continue logging cases.',
            'trial_expired': True,
            'reason': trial_info.get('reason', 'unknown'),
            'cases_used': trial_info.get('cases_used', 0),
            'case_limit': trial_info.get('case_limit', 25),
            'upgrade_url': '/login?tab=register'
        }), 403
    
    data = request.json
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    # Auto-generate case_id if not provided: CCL-YYYY-NNNN
    case_id = data.get('caseId') or data.get('case_id')
    if not case_id:
        from datetime import datetime
        year = datetime.now().year
        cur.execute("SELECT COUNT(*) FROM cases WHERE user_id = %s AND case_id LIKE %s",
                     (request.user['id'], f'CCL-{year}-%'))
        count = cur.fetchone()['count']
        case_id = f'CCL-{year}-{count + 1:04d}'
    cur.execute('''
        INSERT INTO cases (user_id, date, age, sex, rotation, procedure_name, cpt_code, 
                          role, approach, attending, complications, ebl, or_time, notes, case_id)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        RETURNING *
    ''', (
        request.user['id'],
        data.get('date'), data.get('age') or None, data.get('sex'),
        data.get('rotation'), data.get('procedure'), data.get('cpt'),
        data.get('role'), data.get('approach'), data.get('attending'),
        data.get('complications', 'None'), data.get('ebl') or None, data.get('orTime') or None,
        data.get('notes'), case_id
    ))
    case = cur.fetchone()
    conn.commit()
    conn.close()
    if case.get('created_at') and hasattr(case['created_at'], 'isoformat'):
        case['created_at'] = case['created_at'].isoformat()
    if case.get('updated_at') and hasattr(case['updated_at'], 'isoformat'):
        case['updated_at'] = case['updated_at'].isoformat()
    return jsonify(case), 201

@app.route('/api/cases/<int:case_id>', methods=['PUT'])
@auth_required
def update_case(case_id):
    data = request.json
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute('SELECT * FROM cases WHERE id = %s AND user_id = %s', (case_id, request.user['id']))
    case = cur.fetchone()
    if not case:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    
    cur.execute('''
        UPDATE cases SET date=%s, age=%s, sex=%s, rotation=%s, procedure_name=%s, cpt_code=%s,
        role=%s, approach=%s, attending=%s, complications=%s, ebl=%s, or_time=%s, notes=%s,
        case_id=%s, updated_at=NOW()
        WHERE id = %s AND user_id = %s RETURNING *
    ''', (
        data.get('date'), data.get('age') or None, data.get('sex'),
        data.get('rotation'), data.get('procedure'), data.get('cpt'),
        data.get('role'), data.get('approach'), data.get('attending'),
        data.get('complications', 'None'), data.get('ebl') or None, data.get('orTime') or None,
        data.get('notes'), data.get('caseId') or data.get('case_id') or case['case_id'],
        case_id, request.user['id']
    ))
    updated = cur.fetchone()
    conn.commit()
    conn.close()
    if updated.get('created_at') and hasattr(updated['created_at'], 'isoformat'):
        updated['created_at'] = updated['created_at'].isoformat()
    if updated.get('updated_at') and hasattr(updated['updated_at'], 'isoformat'):
        updated['updated_at'] = updated['updated_at'].isoformat()
    return jsonify(updated)

@app.route('/api/cases/<int:case_id>', methods=['DELETE'])
@auth_required
def delete_case(case_id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute('DELETE FROM cases WHERE id = %s AND user_id = %s', (case_id, request.user['id']))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ============ MILESTONES ============

@app.route('/api/milestones', methods=['GET'])
@auth_required
def get_milestones():
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute('SELECT * FROM milestones WHERE user_id = %s', (request.user['id'],))
    milestones = cur.fetchall()
    conn.close()
    return jsonify(milestones)

@app.route('/api/milestones', methods=['POST'])
@auth_required
def update_milestone():
    data = request.json
    conn = get_db()
    cur = conn.cursor()
    cur.execute('''
        INSERT INTO milestones (user_id, category, count) VALUES (%s, %s, %s)
        ON CONFLICT(user_id, category) DO UPDATE SET count = %s
    ''', (request.user['id'], data['category'], data['count'], data['count']))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ============ ATTENDINGS ============

@app.route('/api/attendings', methods=['GET'])
@auth_required
def get_attendings():
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute('SELECT id, name, specialty FROM user_attendings WHERE user_id = %s ORDER BY name ASC', (request.user['id'],))
    attendings = cur.fetchall()
    conn.close()
    return jsonify(attendings)

@app.route('/api/attendings', methods=['POST'])
@auth_required
def add_attending():
    data = request.json
    name = (data.get('name') or '').strip()
    specialty = (data.get('specialty') or '').strip() or None
    if not name:
        return jsonify({'error': 'Name is required'}), 400
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute('INSERT INTO user_attendings (user_id, name, specialty) VALUES (%s, %s, %s) RETURNING id, name, specialty',
                     (request.user['id'], name, specialty))
        attending = cur.fetchone()
        conn.commit()
        conn.close()
        return jsonify(attending), 201
    except psycopg2.errors.UniqueViolation:
        return jsonify({'error': 'Attending already saved'}), 409
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/attendings/<int:attending_id>', methods=['DELETE'])
@auth_required
def delete_attending(attending_id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute('DELETE FROM user_attendings WHERE id = %s AND user_id = %s', (attending_id, request.user['id']))
    deleted = cur.rowcount
    conn.commit()
    conn.close()
    if deleted == 0:
        return jsonify({'error': 'Not found'}), 404
    return jsonify({'success': True})

# ============ STATS ============

@app.route('/api/stats', methods=['GET'])
@auth_required
def get_stats():
    conn = get_db()
    cur = conn.cursor()
    uid = request.user['id']
    
    cur.execute('SELECT COUNT(*) FROM cases WHERE user_id=%s', (uid,))
    total = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM cases WHERE user_id=%s AND role='Primary Surgeon'", (uid,))
    primary = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM cases WHERE user_id=%s AND role='First Assist'", (uid,))
    assist = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM cases WHERE user_id=%s AND complications != 'None' AND complications IS NOT NULL", (uid,))
    comps = cur.fetchone()[0]
    cur.execute("SELECT AVG(or_time) FROM cases WHERE user_id=%s AND or_time > 0", (uid,))
    avg_or = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM cases WHERE user_id=%s AND approach='Laparoscopic'", (uid,))
    lap = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM cases WHERE user_id=%s AND approach='Robotic'", (uid,))
    robotic = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM cases WHERE user_id=%s AND approach='Open'", (uid,))
    open_cases = cur.fetchone()[0]
    
    cur2 = conn.cursor(cursor_factory=RealDictCursor)
    cur2.execute("SELECT rotation, COUNT(*) as count FROM cases WHERE user_id=%s GROUP BY rotation ORDER BY count DESC", (uid,))
    rotations = cur2.fetchall()
    
    conn.close()
    
    return jsonify({
        'total': total, 'primary': primary, 'assist': assist,
        'complications': comps,
        'compRate': round((comps/total)*100, 1) if total > 0 else 0,
        'avgOR': round(float(avg_or)) if avg_or else 0,
        'laparoscopic': lap, 'robotic': robotic, 'open': open_cases,
        'rotations': rotations
    })

# ============ USER PROGRAM TYPE ============

@app.route('/api/user/program', methods=['PUT'])
@auth_required
def update_program_type():
    data = request.json
    program_type = data.get('program_type', 'medical')
    if program_type not in ('medical', 'rnfa', 'pa', 'np', 'resident'):
        return jsonify({'error': 'Invalid program type'}), 400
    conn = get_db()
    cur = conn.cursor()
    cur.execute('UPDATE users SET program_type = %s WHERE id = %s', (program_type, request.user['id']))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'program_type': program_type})

# ============ RNFA CASES ============

@app.route('/api/rnfa-cases', methods=['GET'])
@auth_required
def get_rnfa_cases():
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute('SELECT * FROM rnfa_cases WHERE user_id = %s ORDER BY case_number DESC, id DESC', (request.user['id'],))
    cases = cur.fetchall()
    conn.close()
    for c in cases:
        if c.get('created_at') and hasattr(c['created_at'], 'isoformat'):
            c['created_at'] = c['created_at'].isoformat()
    return jsonify(cases)

@app.route('/api/rnfa-cases', methods=['POST'])
@auth_required
def add_rnfa_case():
    # TRIAL ENFORCEMENT — check before allowing case creation
    can_log, trial_info = check_trial_status(request.user)
    if not can_log:
        return jsonify({
            'error': 'Trial limit reached. Upgrade to continue logging cases.',
            'trial_expired': True,
            'reason': trial_info.get('reason', 'unknown'),
            'cases_used': trial_info.get('cases_used', 0),
            'case_limit': trial_info.get('case_limit', 25),
            'upgrade_url': '/login?tab=register'
        }), 403
    
    data = request.json
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    # Auto-increment case number per user
    cur.execute('SELECT COALESCE(MAX(case_number), 0) + 1 AS next_num FROM rnfa_cases WHERE user_id = %s', (request.user['id'],))
    next_num = cur.fetchone()['next_num']
    cur.execute('''
        INSERT INTO rnfa_cases (user_id, case_number, procedure_name, surgeon_name, case_date, case_time,
                               preop_time, intraop_time, postop_time, preop_behaviors, draping_prepping,
                               special_instruments, medical_devices, exposure_hemostasis, layer_closure,
                               postop_behaviors, learning_log)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        RETURNING *
    ''', (
        request.user['id'], next_num,
        data.get('procedure_name'), data.get('surgeon_name'),
        data.get('case_date'), data.get('case_time'),
        data.get('preop_time') or 0, data.get('intraop_time') or 0, data.get('postop_time') or 0,
        data.get('preop_behaviors'), data.get('draping_prepping'),
        data.get('special_instruments'), data.get('medical_devices'),
        data.get('exposure_hemostasis'), data.get('layer_closure'),
        data.get('postop_behaviors'), data.get('learning_log')
    ))
    case = cur.fetchone()
    conn.commit()
    conn.close()
    if case.get('created_at') and hasattr(case['created_at'], 'isoformat'):
        case['created_at'] = case['created_at'].isoformat()
    return jsonify(case), 201

@app.route('/api/rnfa-cases/<int:case_id>', methods=['PUT'])
@auth_required
def update_rnfa_case(case_id):
    data = request.json
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute('SELECT * FROM rnfa_cases WHERE id = %s AND user_id = %s', (case_id, request.user['id']))
    case = cur.fetchone()
    if not case:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    cur.execute('''
        UPDATE rnfa_cases SET procedure_name=%s, surgeon_name=%s, case_date=%s, case_time=%s,
        preop_time=%s, intraop_time=%s, postop_time=%s, preop_behaviors=%s, draping_prepping=%s,
        special_instruments=%s, medical_devices=%s, exposure_hemostasis=%s, layer_closure=%s,
        postop_behaviors=%s, learning_log=%s
        WHERE id = %s AND user_id = %s RETURNING *
    ''', (
        data.get('procedure_name'), data.get('surgeon_name'),
        data.get('case_date'), data.get('case_time'),
        data.get('preop_time') or 0, data.get('intraop_time') or 0, data.get('postop_time') or 0,
        data.get('preop_behaviors'), data.get('draping_prepping'),
        data.get('special_instruments'), data.get('medical_devices'),
        data.get('exposure_hemostasis'), data.get('layer_closure'),
        data.get('postop_behaviors'), data.get('learning_log'),
        case_id, request.user['id']
    ))
    updated = cur.fetchone()
    conn.commit()
    conn.close()
    if updated.get('created_at') and hasattr(updated['created_at'], 'isoformat'):
        updated['created_at'] = updated['created_at'].isoformat()
    return jsonify(updated)

@app.route('/api/rnfa-cases/<int:case_id>', methods=['DELETE'])
@auth_required
def delete_rnfa_case(case_id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute('DELETE FROM rnfa_cases WHERE id = %s AND user_id = %s', (case_id, request.user['id']))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/rnfa-cases/export', methods=['GET'])
@auth_required
def export_rnfa_cases():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return jsonify({'error': 'openpyxl not installed on server'}), 500

    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute('SELECT * FROM rnfa_cases WHERE user_id = %s ORDER BY case_number ASC', (request.user['id'],))
    cases = cur.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "RNFA Procedure Log"

    # Styles
    title_font = Font(name='Arial', size=14, bold=True, color='1B2A4A')
    subtitle_font = Font(name='Arial', size=11, italic=True, color='555555')
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
    data_font = Font(name='Arial', size=10)
    wrap_align = Alignment(wrap_text=True, vertical='top')
    center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Row 1: Title
    ws.merge_cells('A1:H1')
    ws['A1'] = 'RNFA CLINICAL PROCEDURE LOG \u2013 RNFA Program'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    # Row 2: Subtitle
    ws.merge_cells('A2:H2')
    ws['A2'] = 'Procedure Log 1 (minimum 65 intraoperative hours)'
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = Alignment(horizontal='center')

    # Row 3: Column headers
    headers = ['Case #', 'Procedure and Surgeon\'s Name', 'Date/time', 'Pre-op time', 'Intra-op time', 'Post-op time', 'First Assisting Behaviors', 'Learning Log']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 60
    ws.column_dimensions['H'].width = 40

    # Data rows
    total_intraop = 0
    for row_idx, c in enumerate(cases, 4):
        # Case #
        ws.cell(row=row_idx, column=1, value=c['case_number']).font = data_font
        ws.cell(row=row_idx, column=1).alignment = center_align
        ws.cell(row=row_idx, column=1).border = thin_border

        # Procedure and Surgeon's Name combined
        proc_surgeon = (c['procedure_name'] or '')
        if c.get('surgeon_name'):
            proc_surgeon += '\nSurgeon: ' + c['surgeon_name']
        cell_b = ws.cell(row=row_idx, column=2, value=proc_surgeon)
        cell_b.font = data_font
        cell_b.alignment = wrap_align
        cell_b.border = thin_border

        # Date/time
        dt_str = (c['case_date'] or '')
        if c.get('case_time'):
            dt_str += ' ' + c['case_time']
        ws.cell(row=row_idx, column=3, value=dt_str).font = data_font
        ws.cell(row=row_idx, column=3).alignment = center_align
        ws.cell(row=row_idx, column=3).border = thin_border

        # Pre-op time
        ws.cell(row=row_idx, column=4, value=c.get('preop_time') or 0).font = data_font
        ws.cell(row=row_idx, column=4).alignment = center_align
        ws.cell(row=row_idx, column=4).border = thin_border

        # Intra-op time
        intraop = c.get('intraop_time') or 0
        total_intraop += intraop
        ws.cell(row=row_idx, column=5, value=intraop).font = data_font
        ws.cell(row=row_idx, column=5).alignment = center_align
        ws.cell(row=row_idx, column=5).border = thin_border

        # Post-op time
        ws.cell(row=row_idx, column=6, value=c.get('postop_time') or 0).font = data_font
        ws.cell(row=row_idx, column=6).alignment = center_align
        ws.cell(row=row_idx, column=6).border = thin_border

        # First Assisting Behaviors (combined)
        behaviors_parts = []
        if c.get('preop_behaviors'):
            behaviors_parts.append('PRE-OP:\n' + c['preop_behaviors'])
        if c.get('draping_prepping'):
            behaviors_parts.append('DRAPING/PREPPING/POSITIONING:\n' + c['draping_prepping'])
        if c.get('special_instruments'):
            behaviors_parts.append('SPECIAL INSTRUMENTS:\n' + c['special_instruments'])
        if c.get('medical_devices'):
            behaviors_parts.append('MEDICAL DEVICES:\n' + c['medical_devices'])
        if c.get('exposure_hemostasis'):
            behaviors_parts.append('EXPOSURE, HEMOSTASIS, TISSUE HANDLING/CUTTING:\n' + c['exposure_hemostasis'])
        if c.get('layer_closure'):
            behaviors_parts.append('LAYER CLOSURE:\n' + c['layer_closure'])
        if c.get('postop_behaviors'):
            behaviors_parts.append('POST-OP:\n' + c['postop_behaviors'])
        behaviors_combined = '\n\n'.join(behaviors_parts)
        cell_g = ws.cell(row=row_idx, column=7, value=behaviors_combined)
        cell_g.font = data_font
        cell_g.alignment = wrap_align
        cell_g.border = thin_border

        # Learning Log
        cell_h = ws.cell(row=row_idx, column=8, value=c.get('learning_log') or '')
        cell_h.font = data_font
        cell_h.alignment = wrap_align
        cell_h.border = thin_border

    # Total row
    total_row = 4 + len(cases)
    ws.merge_cells(f'A{total_row}:D{total_row}')
    total_cell = ws.cell(row=total_row, column=1, value='Total Intra-op Hours:')
    total_cell.font = Font(name='Arial', size=11, bold=True, color='1B2A4A')
    total_cell.alignment = Alignment(horizontal='right')
    total_cell.border = thin_border
    for col in range(2, 5):
        ws.cell(row=total_row, column=col).border = thin_border

    hours_cell = ws.cell(row=total_row, column=5, value=round(total_intraop / 60.0, 1))
    hours_cell.font = Font(name='Arial', size=11, bold=True, color='2E8B8B')
    hours_cell.alignment = center_align
    hours_cell.border = thin_border

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    user_name = request.user.get('name', 'RNFA_Student').replace(' ', '_')
    filename = f'RNFA_Procedure_Log_{user_name}.xlsx'
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )

@app.route('/api/rnfa-stats', methods=['GET'])
@auth_required
def get_rnfa_stats():
    conn = get_db()
    cur = conn.cursor()
    uid = request.user['id']
    cur.execute('SELECT COUNT(*) FROM rnfa_cases WHERE user_id=%s', (uid,))
    total = cur.fetchone()[0]
    cur.execute('SELECT COALESCE(SUM(intraop_time), 0) FROM rnfa_cases WHERE user_id=%s', (uid,))
    total_intraop = cur.fetchone()[0]
    cur.execute('SELECT COALESCE(SUM(preop_time), 0) FROM rnfa_cases WHERE user_id=%s', (uid,))
    total_preop = cur.fetchone()[0]
    cur.execute('SELECT COALESCE(SUM(postop_time), 0) FROM rnfa_cases WHERE user_id=%s', (uid,))
    total_postop = cur.fetchone()[0]
    conn.close()
    return jsonify({
        'total_cases': total,
        'total_intraop_minutes': total_intraop,
        'total_intraop_hours': round(total_intraop / 60.0, 1),
        'total_preop_minutes': total_preop,
        'total_preop_hours': round(total_preop / 60.0, 1),
        'total_postop_minutes': total_postop,
        'total_postop_hours': round(total_postop / 60.0, 1),
        'intraop_goal_hours': 65,
        'intraop_progress_pct': round(min((total_intraop / 60.0) / 65.0 * 100, 100), 1)
    })

# ============ ADMIN PANEL ============

ADMIN_SECRET = 'stallard2026admin'

@app.route('/admin')
def admin_panel():
    key = request.args.get('key', '')
    if key != ADMIN_SECRET:
        return '<h1>Access Denied</h1>', 403
    return send_from_directory('.', 'admin.html')

@app.route('/api/admin/stats', methods=['GET'])
def admin_stats():
    key = request.args.get('key', '')
    if key != ADMIN_SECRET:
        return jsonify({'error': 'Access denied'}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Total registered users
        cur.execute('SELECT COUNT(*) FROM users')
        total_users = cur.fetchone()[0]
        
        # Active users (users with cases in last 7 days)
        cur.execute('''
            SELECT COUNT(DISTINCT user_id) FROM cases 
            WHERE created_at >= NOW() - INTERVAL '7 days'
        ''')
        active_users = cur.fetchone()[0]
        
        # Total cases (medical + RNFA)
        cur.execute('SELECT COUNT(*) FROM cases')
        medical_cases = cur.fetchone()[0]
        cur.execute('SELECT COUNT(*) FROM rnfa_cases')
        rnfa_cases = cur.fetchone()[0]
        total_cases = medical_cases + rnfa_cases
        
        # Plan breakdown
        cur.execute('''
            SELECT 
                plan,
                COUNT(*) as count,
                CASE 
                    WHEN plan = 'annual' AND plan_expires_at IS NOT NULL AND plan_expires_at < NOW()::text THEN 'expired'
                    WHEN plan = '90day' AND plan_expires_at IS NOT NULL AND plan_expires_at < NOW()::text THEN 'expired'
                    ELSE plan
                END as effective_plan
            FROM users 
            GROUP BY plan, effective_plan
            ORDER BY count DESC
        ''')
        plan_data = cur.fetchall()
        
        # Consolidate plan counts
        plan_counts = {}
        for row in plan_data:
            plan = row[2]  # effective_plan
            count = row[1]
            if plan in plan_counts:
                plan_counts[plan] += count
            else:
                plan_counts[plan] = count
        
        paid_users = plan_counts.get('annual', 0) + plan_counts.get('subscription', 0)
        trial_users = plan_counts.get('trial', 0)
        expired_users = plan_counts.get('expired', 0)
        
        conn.close()
        
        return jsonify({
            'total_users': total_users,
            'active_users': active_users,
            'total_cases': total_cases,
            'medical_cases': medical_cases,
            'rnfa_cases': rnfa_cases,
            'paid_users': paid_users,
            'trial_users': trial_users,
            'expired_users': expired_users,
            'plan_breakdown': plan_counts
        })
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/admin/users', methods=['GET'])
def admin_users():
    key = request.args.get('key', '')
    if key != ADMIN_SECRET:
        return jsonify({'error': 'Access denied'}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Get all users with case counts
        cur.execute('''
            SELECT 
                u.id, u.email, u.name, u.plan, u.program_type, u.plan_expires_at, u.created_at,
                COALESCE(c.case_count, 0) as medical_cases,
                COALESCE(r.rnfa_count, 0) as rnfa_cases,
                COALESCE(c.case_count, 0) + COALESCE(r.rnfa_count, 0) as total_cases
            FROM users u
            LEFT JOIN (
                SELECT user_id, COUNT(*) as case_count 
                FROM cases 
                GROUP BY user_id
            ) c ON u.id = c.user_id
            LEFT JOIN (
                SELECT user_id, COUNT(*) as rnfa_count 
                FROM rnfa_cases 
                GROUP BY user_id
            ) r ON u.id = r.user_id
            ORDER BY u.created_at DESC
        ''')
        users = cur.fetchall()
        
        # Convert timestamps to strings
        for user in users:
            if user.get('created_at') and hasattr(user['created_at'], 'isoformat'):
                user['created_at'] = user['created_at'].isoformat()
        
        conn.close()
        return jsonify(users)
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/admin/activity', methods=['GET'])
def admin_activity():
    key = request.args.get('key', '')
    if key != ADMIN_SECRET:
        return jsonify({'error': 'Access denied'}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Daily signups for last 30 days
        cur.execute('''
            SELECT 
                DATE(created_at) as signup_date,
                COUNT(*) as signups
            FROM users 
            WHERE created_at >= NOW() - INTERVAL '30 days'
            GROUP BY DATE(created_at)
            ORDER BY signup_date ASC
        ''')
        signup_data = cur.fetchall()
        
        # Daily case counts for last 30 days
        cur.execute('''
            SELECT 
                DATE(created_at) as case_date,
                COUNT(*) as medical_cases
            FROM cases 
            WHERE created_at >= NOW() - INTERVAL '30 days'
            GROUP BY DATE(created_at)
            ORDER BY case_date ASC
        ''')
        medical_case_data = cur.fetchall()
        
        cur.execute('''
            SELECT 
                DATE(created_at) as case_date,
                COUNT(*) as rnfa_cases
            FROM rnfa_cases 
            WHERE created_at >= NOW() - INTERVAL '30 days'
            GROUP BY DATE(created_at)
            ORDER BY case_date ASC
        ''')
        rnfa_case_data = cur.fetchall()
        
        # Top 10 most active users
        cur.execute('''
            SELECT 
                u.name, u.email,
                COALESCE(c.case_count, 0) + COALESCE(r.rnfa_count, 0) as total_cases
            FROM users u
            LEFT JOIN (
                SELECT user_id, COUNT(*) as case_count 
                FROM cases 
                GROUP BY user_id
            ) c ON u.id = c.user_id
            LEFT JOIN (
                SELECT user_id, COUNT(*) as rnfa_count 
                FROM rnfa_cases 
                GROUP BY user_id
            ) r ON u.id = r.user_id
            ORDER BY total_cases DESC
            LIMIT 10
        ''')
        top_users = cur.fetchall()
        
        conn.close()
        
        return jsonify({
            'signups': [{'date': str(row[0]), 'count': row[1]} for row in signup_data],
            'medical_cases': [{'date': str(row[0]), 'count': row[1]} for row in medical_case_data],
            'rnfa_cases': [{'date': str(row[0]), 'count': row[1]} for row in rnfa_case_data],
            'top_users': [{'name': row[0], 'email': row[1], 'cases': row[2]} for row in top_users]
        })
    except Exception as e:
        return jsonify({'error': str(e)})

# ============ STATIC FILES ============

@app.route('/googlef8c23c2804f8d2c1.html')
def google_verify():
    return send_from_directory('.', 'googlef8c23c2804f8d2c1.html')

@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/preview')
def preview():
    return send_from_directory('.', 'index.html')

@app.route('/demo')
def demo():
    return send_from_directory('.', 'demo.html')

@app.route('/login')
@app.route('/signin')
@app.route('/register')
def login_page():
    return send_from_directory('.', 'login.html')

@app.route('/prep')
def prep():
    return send_from_directory('.', 'prep.html')

@app.route('/legal')
def legal():
    return send_from_directory('.', 'legal.html')

@app.route('/privacy')
def privacy():
    return send_from_directory('.', 'privacy.html')

@app.route('/api/preview-login', methods=['POST'])
def preview_login():
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    email = 'admin@clinicalcaselog.com'
    cur.execute('SELECT * FROM users WHERE email = %s', (email,))
    user = cur.fetchone()
    
    if not user:
        token = secrets.token_urlsafe(32)
        pw_hash = hash_password('preview123')
        cur.execute('INSERT INTO users (email, password_hash, name, token, plan) VALUES (%s, %s, %s, %s, %s)',
                     (email, pw_hash, 'Admin Preview', token, 'subscription'))
        conn.commit()
    else:
        token = user['token']
    
    conn.close()
    return jsonify({'token': token, 'name': 'Admin Preview', 'email': email, 'plan': 'subscription'})



# ==========================================
# ANALYTICS - Page Visit Tracking
# ==========================================

def init_analytics():
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS page_visits (
                id SERIAL PRIMARY KEY,
                page TEXT NOT NULL,
                referrer TEXT,
                user_agent TEXT,
                ip_address TEXT,
                visited_at TIMESTAMP DEFAULT NOW()
            )
        """)
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        print(f"Analytics init error: {e}")

@app.route('/api/track', methods=['POST'])
def track_visit():
    try:
        data = request.get_json() or {}
        page = data.get('page', 'unknown')
        referrer = data.get('referrer', request.referrer or '')
        user_agent = request.headers.get('User-Agent', '')[:500]
        ip = request.headers.get('X-Forwarded-For', request.remote_addr or '')
        conn = get_db()
        cur = conn.cursor()
        cur.execute('INSERT INTO page_visits (page, referrer, user_agent, ip_address) VALUES (%s, %s, %s, %s)',
            (page, referrer[:500], user_agent, ip[:100]))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/analytics')
def get_analytics():
    key = request.args.get('key', '')
    if key != 'stallard2026admin':
        return jsonify({'error': 'Unauthorized'}), 401
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT page, COUNT(*) as visits, COUNT(DISTINCT ip_address) as unique_visitors, MAX(visited_at) as last_visit FROM page_visits GROUP BY page ORDER BY visits DESC")
        by_page = cur.fetchall()
        cur.execute("SELECT page, COUNT(*) as visits FROM page_visits WHERE visited_at >= CURRENT_DATE GROUP BY page")
        today = cur.fetchall()
        cur.execute("SELECT DATE(visited_at) as date, page, COUNT(*) as visits FROM page_visits WHERE visited_at >= CURRENT_DATE - INTERVAL '30 days' GROUP BY DATE(visited_at), page ORDER BY date DESC")
        by_day = cur.fetchall()
        cur.execute("SELECT referrer, COUNT(*) as visits FROM page_visits WHERE referrer != '' AND referrer IS NOT NULL GROUP BY referrer ORDER BY visits DESC LIMIT 20")
        referrers = cur.fetchall()
        cur.close()
        conn.close()
        return jsonify({'by_page': by_page, 'today': today, 'by_day': by_day, 'referrers': referrers})
    except Exception as e:
        return jsonify({'error': str(e)})

init_analytics()

# ============ STRIPE PAYMENTS ============
import stripe as stripe_lib
stripe_lib.api_key = os.environ.get("STRIPE_SECRET_KEY", "")

@app.route('/api/check-subscription', methods=['POST'])
def check_subscription():
    """Check if an email has an active subscription via Stripe (primary) or Gumroad (fallback)."""
    data = request.get_json() or {}
    email = (data.get('email') or '').strip().lower()
    if not email:
        return jsonify({'active': False, 'source': None, 'error': 'Email required'}), 400

    # 1) Check Stripe first
    try:
        customers = stripe_lib.Customer.list(email=email, limit=5)
        for customer in customers.auto_paging_iter():
            subs = stripe_lib.Subscription.list(customer=customer.id, status='active', limit=1)
            if subs.data:
                return jsonify({'active': True, 'source': 'stripe', 'plan': 'subscription'})
            subs_trial = stripe_lib.Subscription.list(customer=customer.id, status='trialing', limit=1)
            if subs_trial.data:
                return jsonify({'active': True, 'source': 'stripe', 'plan': 'subscription'})
    except Exception as e:
        print(f"Stripe check error: {e}")

    # 2) Fallback to Gumroad for legacy purchases
    import requests as req
    GUMROAD_TOKEN = "X_X3q6Cw7bJuRkvmwKRGIwhN9Comr3wkjKkMXrv-ZFE"
    try:
        r = req.get(f'https://api.gumroad.com/v2/sales?access_token={GUMROAD_TOKEN}&email={email}', timeout=15)
        if r.status_code == 200:
            sales = r.json().get('sales', [])
            for sale in sales:
                if sale.get('refunded') or sale.get('disputed'):
                    continue
                return jsonify({'active': True, 'source': 'gumroad', 'plan': 'annual'})
    except Exception as e:
        print(f"Gumroad check error: {e}")

    return jsonify({'active': False, 'source': None})

@app.route('/api/create-checkout-session', methods=['POST'])
def create_checkout_session():
    try:
        data = request.get_json() or {}
        user_email = data.get('email', '')
        
        session = stripe_lib.checkout.Session.create(
            payment_method_types=['card'],
            line_items=[{
                'price': os.environ.get('STRIPE_PRICE_ID', 'price_1TDsP2C0LC9x3wbewDKfgfXh'),
                'quantity': 1,
            }],
            mode='subscription',
            success_url='https://clinicalcaselog.com/login?paid=true&tab=register',
            cancel_url='https://clinicalcaselog.com/login?cancelled=true',
            customer_email=user_email if user_email else None,
            metadata={
                'source': 'clinical-case-log'
            }
        )
        return jsonify({'url': session.url})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/stripe-webhook', methods=['POST'])
def stripe_webhook():
    payload = request.get_data(as_text=True)
    
    try:
        event = stripe_lib.Event.construct_from(
            json.loads(payload), stripe_lib.api_key
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    
    if event.type == 'checkout.session.completed':
        session_obj = event.data.object
        customer_email = session_obj.get('customer_email') or session_obj.get('customer_details', {}).get('email')
        if customer_email:
            try:
                conn = get_db()
                cur = conn.cursor()
                sub_id = session_obj.get('subscription')
                expires = None
                if sub_id:
                    try:
                        sub = stripe_lib.Subscription.retrieve(sub_id)
                        expires = datetime.utcfromtimestamp(sub.current_period_end).strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        pass
                cur.execute("UPDATE users SET plan = 'subscription', plan_expires_at = %s WHERE email = %s", (expires, customer_email))
                conn.commit()
                conn.close()
                print(f"✅ Stripe webhook: Activated subscription for {customer_email}")
            except Exception as e:
                print(f"❌ Stripe webhook DB error: {e}")
    
    return jsonify({'status': 'ok'})

if __name__ == '__main__':
    print(f"🚀 Server starting on port 8080...")
    app.run(host='0.0.0.0', port=8080, debug=False)
